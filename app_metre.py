import streamlit as st
import fitz  # PyMuPDF
import re
import pandas as pd
import json
import requests
import datetime
from io import BytesIO
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pytesseract
from PIL import Image

st.set_page_config(page_title="METRE-TEST System", page_icon="🏗️", layout="wide")

# ==========================================
# INJECTION CSS POUR DESIGN PRO
# ==========================================
st.markdown("""
<style>
/* Masquer le menu et le footer de Streamlit pour faire plus "Logiciel" */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}

/* Fond global de l'application en clair */
.stApp {
    background-color: #f4f6f9;
}

/* Design du Header / Logo - Thème Clair */
.header-container {
    background-color: #ffffff;
    padding: 20px;
    border-radius: 10px;
    margin-bottom: 30px;
    display: flex;
    align-items: center;
    border-left: 8px solid #3498db;
    box-shadow: 0 4px 10px rgba(0,0,0,0.05);
}
.logo-icon {
    font-size: 45px;
    margin-right: 20px;
}
.app-title {
    color: #2c3e50;
    font-size: 30px;
    font-weight: 800;
    margin: 0;
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
}
.app-subtitle {
    color: #7f8c8d;
    font-size: 15px;
    margin: 0;
    margin-top: 5px;
}
</style>
""", unsafe_allow_html=True)

# ==========================================
# 0. API KEY ET BASE DE DONNÉES
# ==========================================
GROQ_TOKEN = "gsk_" + "cVs6luXGdAoyHqRSG3BVWGdyb3FYK4P0tNs1a2g1izU6K87AhGTk"
API_URL = "https://api.groq.com/openai/v1/chat/completions"

# CATALOGUE DES POIDS DES PROFILÉS MÉTALLIQUES (en Kg / ml) - Basé sur ArcelorMittal
CATALOGUE_PROFILS = {
    "IPE80": 6.0,
    "IPE100": 8.1,
    "IPE120": 10.4,
    "IPE140": 12.9,
    "IPE160": 15.8,
    "IPE180": 18.8,
    "IPE200": 22.4,
    "IPE220": 26.2,
    "IPE240": 30.7,
    "IPE270": 36.1,
    "IPE300": 42.2,
    "IPE330": 49.1,
    "IPE360": 57.1,
    "IPE400": 66.3,
    "IPE450": 77.6,
    "IPE500": 90.7,
    "IPE550": 106.0,
    "IPE600": 122.0,
    "IPN80": 5.94,
    "IPN100": 8.34,
    "IPN120": 11.1,
    "IPN140": 14.3,
    "IPN160": 17.9,
    "IPN180": 21.9,
    "IPN200": 26.2,
    "IPN220": 31.1,
    "IPN240": 36.2,
    "IPN260": 41.9,
    "IPN280": 47.9,
    "IPN300": 54.2,
    "IPN320": 61.0,
    "IPN340": 68.0,
    "IPN360": 76.1,
    "IPN380": 84.0,
    "IPN400": 92.4,
    "IPN450": 115.0,
    "IPN500": 141.0,
    "IPN550": 166.0,
    "HEA100": 16.7,
    "HEA120": 19.9,
    "HEA140": 24.7,
    "HEA160": 30.4,
    "HEA180": 35.5,
    "HEA200": 42.3,
    "HEA220": 50.5,
    "HEA240": 60.3,
    "HEA260": 68.2,
    "HEA280": 76.4,
    "HEA300": 88.3,
    "HEA320": 97.6,
    "HEA340": 105.0,
    "HEA360": 112.0,
    "HEA400": 125.0,
    "HEA450": 140.0,
    "HEA500": 155.0,
    "HEA550": 166.0,
    "HEA600": 178.0,
    "HEA650": 190.0,
    "HEA700": 204.0,
    "HEA800": 224.0,
    "HEA900": 252.0,
    "HEA1000": 272.0,
    "HEB100": 20.4,
    "HEB120": 26.7,
    "HEB140": 33.7,
    "HEB160": 42.6,
    "HEB180": 51.2,
    "HEB200": 61.3,
    "HEB220": 71.5,
    "HEB240": 83.2,
    "HEB260": 93.0,
    "HEB280": 103.0,
    "HEB300": 117.0,
    "HEB320": 127.0,
    "HEB340": 134.0,
    "HEB360": 142.0,
    "HEB400": 155.0,
    "HEB450": 171.0,
    "HEB500": 187.0,
    "HEB550": 199.0,
    "HEB600": 212.0,
    "HEB650": 225.0,
    "HEB700": 241.0,
    "HEB800": 262.0,
    "HEB900": 291.0,
    "HEB1000": 314.0,
    "UPN80": 8.65,
    "UPN100": 10.6,
    "UPN120": 13.4,
    "UPN140": 16.0,
    "UPN160": 18.8,
    "UPN180": 22.0,
    "UPN200": 25.3,
    "UPN220": 29.4,
    "UPN240": 33.2,
    "UPN260": 37.9,
    "UPN280": 41.8,
    "UPN300": 46.2,
    "UPN320": 59.5,
    "UPN350": 60.6,
    "UPN380": 63.1,
    "UPN400": 71.8,
    "UAP80": 8.38,
    "UAP100": 10.5,
    "UAP130": 13.7,
    "UAP150": 17.9,
    "UAP175": 21.2,
    "UAP200": 25.1,
    "UAP220": 28.5,
    "UAP250": 34.4,
    "UAP300": 46.0,
}

BASE_DONNEES = {
    "L70*7": {"desc": "Cornière à ailes égales 70x7", "unite": "ml", "poids_u": 7.38},
    "BOULON M16": {"desc": "Boulon d'assemblage M16 HR", "unite": "U", "poids_u": 0.15},
    "PL 300*300*20": {"desc": "Platine d'ancrage 300x300 Ep:20mm", "unite": "U", "poids_u": 14.13},
    "SIKAGROUT": {"desc": "Mortier de scellement Sikagrout", "unite": "Sac", "poids_u": 25.0},
    "POTEAU BETON": {"desc": "Poteau en Béton Armé", "unite": "m3", "poids_u": 2500.0},
    "TUBE EN PVC": {"desc": "Tube PVC Évacuation", "unite": "ml", "poids_u": 1.5},
    "BARDAGE": {"desc": "Revêtement / Bardage", "unite": "m²", "poids_u": 10.0},
}

import re
import math

def get_item_info(item_name):
    item_upper = item_name.upper().replace(" ", "")
    
    # 1. Recherche dans le catalogue ArcelorMittal (IPE, HEA, etc.)
    for profil_key, poids_ml in CATALOGUE_PROFILS.items():
        if profil_key in item_upper:
            return {"desc": f"Profilé {profil_key} (Acier S275)", "unite": "ml", "poids_u": poids_ml}
            
    # 2. Calculateur Automatique (Intelligence) pour Profilés Formés à Froid & Plats
    # Densité de l'acier = 7850 kg/m³ -> 0.00785 kg/mm²/m
    densite = 0.00785
    
    # -> A) CORNIÈRES (ex: L50x5, L70*7, Corniere 50x50x5)
    match_l = re.search(r'(?:CORNIERE|CORNIÈRE|L)\s*(\d+)(?:X|\*)(\d+)', item_upper)
    if match_l:
        a = float(match_l.group(1))
        e = float(match_l.group(2))
        poids_calcule = ((2 * a - e) * e) * densite
        return {"desc": f"Cornière à ailes égales {int(a)}x{int(e)}", "unite": "ml", "poids_u": round(poids_calcule, 2)}
        
    # -> B) PLATS (ex: PLAT 100x10)
    match_plat = re.search(r'PLAT\s*(\d+)(?:X|\*)(\d+)', item_upper)
    if match_plat:
        larg = float(match_plat.group(1))
        ep = float(match_plat.group(2))
        return {"desc": f"Plat Acier {int(larg)}x{int(ep)}", "unite": "ml", "poids_u": round(larg * ep * densite, 2)}
        
    # -> C) TUBES CARRÉS / RECTANGULAIRES (ex: TUBE 100x100x4)
    match_tube_rect = re.search(r'TUBE(?:.*?)(\d+)(?:X|\*)(\d+)(?:X|\*)(\d+)', item_upper)
    if match_tube_rect:
        a = float(match_tube_rect.group(1))
        b = float(match_tube_rect.group(2))
        e = float(match_tube_rect.group(3))
        # Poids approx pour tube rectangulaire (périmètre moyen * épaisseur)
        poids_calcule = (2 * (a + b) - 4 * e) * e * densite
        return {"desc": f"Tube Rectangulaire/Carré {int(a)}x{int(b)} ép:{int(e)}", "unite": "ml", "poids_u": round(poids_calcule, 2)}
        
    # -> D) TUBES RONDS (ex: TUBE Ø114.3x3.2)
    match_tube_rond = re.search(r'TUBE(?:.*?)(\d+(?:\.\d+)?)(?:X|\*)(\d+(?:\.\d+)?)', item_upper)
    if match_tube_rond:
        d = float(match_tube_rond.group(1))
        e = float(match_tube_rond.group(2))
        poids_calcule = math.pi * (d - e) * e * densite
        return {"desc": f"Tube Rond Ø{d} ép:{e}", "unite": "ml", "poids_u": round(poids_calcule, 2)}
        
    # -> E) TÔLES NOIRES / PLATINES / RAIDISSEURS (ex: TN300*300*20)
    match_tn = re.search(r'(?:TN|PLAQUE|PLATINE|RAIDISSEUR)(?:.*?)(\d+)(?:X|\*)(\d+)(?:X|\*)(\d+)', item_upper)
    if match_tn:
        a = float(match_tn.group(1))
        b = float(match_tn.group(2))
        e = float(match_tn.group(3))
        # Poids d'une plaque (Volume en m3 * 8000 kg/m3 pour inclure marge soudures/découpe comme dans la pratique)
        poids_calcule = (a / 1000) * (b / 1000) * (e / 1000) * 8000
        return {"desc": f"Platine/Tôle {int(a)}x{int(b)} ép:{int(e)}", "unite": "U", "poids_u": round(poids_calcule, 3)}

    # 3. Base de données classique (Boulons, Platines, Béton, etc.)
    for key in BASE_DONNEES.keys():
        if key.replace(" ", "") in item_upper:
            return BASE_DONNEES[key]
            
    if "BÉTON" in item_upper or "BETON" in item_upper: return {"desc": "Ouvrage en Béton", "unite": "m3", "poids_u": 2500.0}
    if "TUBE" in item_upper or "PVC" in item_upper: return {"desc": f"Tube PVC {item_name}", "unite": "ml", "poids_u": 1.5}
    if "TOLE" in item_upper or "TÔLE" in item_upper or "BARDAGE" in item_upper: return {"desc": f"Tôle / Bardage", "unite": "m²", "poids_u": 10.0}
    if "SIKA" in item_upper: return {"desc": "Produit d'étanchéité/scellement", "unite": "Sac", "poids_u": 25.0}
    if "BOULON" in item_upper or "BLS" in item_upper or "TIGE" in item_upper: return {"desc": f"Fixation {item_name}", "unite": "U", "poids_u": 0.20}
    if "PL" in item_upper or "PLATINE" in item_upper or "GOUSSET" in item_upper: return {"desc": f"Platine / Gousset", "unite": "U", "poids_u": 5.0}
    
    # 4. Fallback pour Profilés Inconnus
    if "IPE" in item_upper or "HEA" in item_upper or "UPN" in item_upper or "HEB" in item_upper: 
        return {"desc": f"Profilé {item_name} (Standard inconnu)", "unite": "ml", "poids_u": 50.0}
        
    return {"desc": f"Élément divers ({item_name})", "unite": "Ens", "poids_u": 1.0}

# ==========================================
# 1. ExtractionEngine (Hybride : Vectoriel + OCR)
# ==========================================
class ExtractionEngine:
    @staticmethod
    def extract_all(doc):
        st.info("📖 **Lecture hybride intelligente (Texte + Vision) :** Le système lit le texte, scanne les images et analyse les lignes de cotation... ⏳")
        full_text = ""
        images_b64 = []
        
        progress_bar = st.progress(0)
        total_pages = len(doc)
        
        for i, page in enumerate(doc):
            # 1. Tentative d'extraction vectorielle
            page_text = page.get_text("text").strip()
            
            # 2. Si le texte est faible ou si la page contient des images (plan mixte), on force l'OCR et Vision
            try:
                pix = page.get_pixmap(dpi=300) # 300 DPI pour Extra Vision / Qualité Maximale
                
                # Sauvegarde en base64 pour la Vision IA (Max 2 pages pour éviter la surcharge)
                if i < 2:
                    import base64
                    # Pour éviter des payloads gigantesques avec 300 DPI, on enregistre en JPEG avec qualité optimisée
                    img = Image.open(io.BytesIO(pix.tobytes("jpeg")))
                    
                    # Optionnel : si l'image dépasse 4000 pixels (Plan A0/A1), on limite sa taille pour ne pas crasher l'API Groq
                    if max(img.size) > 4000:
                        try: resample_filter = Image.Resampling.LANCZOS
                        except AttributeError: resample_filter = Image.LANCZOS
                        img.thumbnail((4000, 4000), resample_filter)
                        
                    buffered = io.BytesIO()
                    img.save(buffered, format="JPEG", quality=85)
                    b64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
                    images_b64.append(b64)
                    
                images = page.get_images(full=True)
                if len(page_text) < 1000 or len(images) > 0:
                    img = Image.open(io.BytesIO(pix.tobytes("jpeg")))
                    ocr_text = pytesseract.image_to_string(img, lang="fra").strip()
                    page_text += "\n" + ocr_text
            except Exception:
                pass
            
            full_text += f"\n--- PAGE {i+1} ---\n" + page_text
            progress_bar.progress((i + 1) / total_pages)
            
        progress_bar.empty()
        return full_text, images_b64

# ==========================================
# 2. Parser Métier (Intégration Llama 3.1 - JSON PRO)
# ==========================================
class ParserMetier:
    @staticmethod
    def parse_with_ai(text, images_b64=None):
        headers = {
            "Authorization": f"Bearer {GROQ_TOKEN}",
            "Content-Type": "application/json"
        }
        # Nettoyage du texte (suppression des espaces multiples) pour économiser des Tokens
        clean_text = re.sub(r'\s+', ' ', text)
        
        prompt = f"""Tu es un expert en BTP et Métré. Analyse le texte suivant extrait d'un plan d'architecture/charpente.
Ta mission est d'extraire 1) Les informations du projet (Cartouche) et 2) TOUS les matériaux.

RÈGLE ABSOLUE : Tu DOIS extraire absolument TOUT ce qui ressemble à un matériau ou élément de construction, même si ce n'est pas standard. Ne laisse RIEN de côté. Si un élément est mentionné plusieurs fois, additionne les quantités.
TRÈS IMPORTANT POUR LA CATÉGORISATION : Tu DOIS déterminer le rôle structural de chaque élément (ex: Poteau, Poutre, Panne, Lisse, Contreventement, Traverse, Platine, Boulonnerie, Divers). Ajoute un champ "role" pour chaque matériau.
TRÈS IMPORTANT POUR L'ACIER ET LES LONGUEURS : Pour les PROFILÉS MÉTALLIQUES (IPE, HEA, HEB, UPN, Tubes, Cornières), l'unité est le mètre linéaire ("ml").
Dans les plans de charpente, la longueur est souvent cachée sous ces formes:
- "L=6.5" ou "L=6500" ou "L: 6500" (en mm souvent)
- "lg: 200" ou "longueur 6m"
- "IPE 400 x 6000" (le x 6000 signifie 6000mm = 6m)
- "8 IPE 400 de 200mm"
TRÈS IMPORTANT POUR LES COTATIONS : Tu as accès aux images du plan ! Regarde attentivement les lignes de cotation (les flèches avec des nombres comme 6000, 4500) dessinées à côté des profilés. Utilise ta vision pour associer la cotation visuelle au profilé !
Tu DOIS IMPÉRATIVEMENT chercher ces indications de longueur pour chaque profilé!
RÈGLE DE CALCUL : 
- Ajoute les champs "nbre_pieces" (entier) et "longueur_unitaire_m" (nombre décimal en mètres).
- Si la longueur n'est pas connue, mets "longueur_unitaire_m": null.
- Multiplie le nombre de pièces par la longueur unitaire en MÈTRES.
Exemple : "4 IPE 200 L=6500" -> Tu mets "nbre_pieces": 4, "longueur_unitaire_m": 6.5, "quantite": 26, "unite": "ml", "infos": "4 pièces de 6.5m".
Si et SEULEMENT SI tu es absolument certain qu'aucune longueur n'est indiquée nulle part pour ce profilé (ni dans le texte, ni sur les cotations de l'image), mets "nbre_pieces": 4, "longueur_unitaire_m": null, "unite": "U" et "infos": "Longueur inconnue". Mais CHERCHE BIEN LA LONGUEUR D'ABORD!

Tu dois répondre UNIQUEMENT avec un objet JSON valide ayant cette structure exacte :
{{
    "metadata": {{
        "projet": "Nom du projet ou titre du plan (laisse vide si introuvable)",
        "societe": "Nom de l'entreprise, maitre d'ouvrage, client ou bureau d'étude (laisse vide si introuvable)",
        "date_plan": "Date trouvée sur le plan (laisse vide si introuvable)",
        "description": "Un bref résumé (1-2 phrases) de ce que représente ce plan (ex: Construction métallique d'un auvent...). Laisse vide si introuvable."
    }},
    "materiaux": [
        {{"role": "Poutre", "element": "IPE 400", "infos": "12 pièces", "nbre_pieces": 12, "longueur_unitaire_m": 0.2, "unite": "ml", "quantite": 2.4}},
        {{"role": "Évacuation", "element": "TUBE EN PVC", "infos": "DN125", "nbre_pieces": 5, "longueur_unitaire_m": null, "unite": "ml", "quantite": 5}}
    ]
}}

Texte à analyser :
{clean_text[:12000]}
"""
        
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt}
                ]
            }
        ]
        
        model_name = "llama-3.1-8b-instant"
        
        if images_b64 and len(images_b64) > 0:
            model_name = "llama-3.2-90b-vision-preview"
            for b64 in images_b64:
                messages[0]["content"].append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/jpeg;base64,{b64}"
                    }
                })
                
        payload = {
            "model": model_name,
            "messages": messages,
            "temperature": 0.1
        }
        
        try:
            response = requests.post(API_URL, headers=headers, json=payload)
            if response.status_code == 200:
                result = response.json()['choices'][0]['message']['content']
                match = re.search(r'(\{.*\}|\[.*\])', result, re.DOTALL)
                if match:
                    try:
                        data_json = json.loads(match.group(0))
                    except:
                        data_json = []
                        
                    if isinstance(data_json, list):
                        items = data_json
                        metadata = {"projet": "", "societe": "", "date_plan": "", "description": ""}
                    elif isinstance(data_json, dict):
                        items = data_json.get("materiaux", data_json.get("matériaux", data_json.get("materials", [])))
                        metadata = data_json.get("metadata", {"projet": "", "societe": "", "date_plan": "", "description": ""})
                    else:
                        items = []
                        metadata = {"projet": "", "societe": "", "date_plan": "", "description": ""}
                    
                    merged = {}
                    for item in items:
                        elem = str(item.get("element", "")).strip().upper()
                        infos = str(item.get("infos", "")).strip()
                        unite = str(item.get("unite", "U")).strip()
                        try: qty = float(item.get("quantite", 1))
                        except: qty = 1.0
                        
                        if elem:
                            key = f"{elem}___{infos}___{unite}"
                            if key in merged:
                                merged[key]["quantite"] += qty
                            else:
                                merged[key] = {"element": elem, "infos": infos, "unite": unite, "quantite": qty}
                                
                    st.success("✔️ **Analyse terminée :** Le plan a été traité et les détails ont été structurés avec succès.")
                    return {"metadata": metadata, "materiaux": list(merged.values()), "raw_response": result}
                else:
                    st.warning("⚠️ Impossible de formater les données. (Passage au mode dégradé).")
                    return ParserMetier.parse_regex(text)
            else:
                st.error(f"Erreur Serveur ({response.status_code}): {response.text} - Passage au mode dégradé.")
                return ParserMetier.parse_regex(text)
        except Exception as e:
            st.error(f"Erreur de connexion : {e}. Passage au mode dégradé.")
            return ParserMetier.parse_regex(text)

    @staticmethod
    def parse_regex(text):
        elements = []
        profiles = re.findall(r'\b(IPE|HEA|HEB|UPN)\s*(\d+)\b', text, re.IGNORECASE)
        for p in profiles: elements.append({"element": f"{p[0].upper()}{p[1]}", "infos": "", "unite": "U", "quantite": 1})
            
        cornieres = re.findall(r'\bL\s*(\d+\*\d+)\b', text, re.IGNORECASE)
        for c in cornieres: elements.append({"element": f"Cornière L{c}", "infos": "", "unite": "U", "quantite": 1})
            
        boulons = re.findall(r'(\d+)\s*Bls\s*(M\d+)', text, re.IGNORECASE)
        for b in boulons:
            elements.append({"element": f"Boulon {b[1].upper()}", "infos": "", "unite": "U", "quantite": int(b[0])})
            
        return {"metadata": {"projet": "Inconnu", "societe": "Inconnu", "date_plan": "", "description": "Extraction manuelle Regex."}, "materiaux": elements}

# ==========================================
# 3. Exporter
# ==========================================
class Exporter:
    @staticmethod
    def to_excel(df, total_general, metadata, logo_bytes=None):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Métré_Détaillé', startrow=8)
            worksheet = writer.sheets['Métré_Détaillé']
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # En-tête PRO avec Métadonnées
            worksheet['A1'] = "MÉTRÉ DÉTAILLÉ (Généré par METRE-TEST)"
            worksheet['A1'].font = Font(bold=True, size=14, color="1F4E78")
            
            worksheet['A3'] = "Projet :"
            worksheet['B3'] = metadata.get('projet', 'Non spécifié')
            worksheet['A4'] = "Société / Client :"
            worksheet['B4'] = metadata.get('societe', 'Non spécifié')
            worksheet['A5'] = "Date du plan :"
            date_plan = metadata.get('date_plan', '')
            worksheet['B5'] = date_plan if date_plan else "Non spécifiée"
            
            worksheet['A6'] = "Généré le :"
            date_export = datetime.datetime.now().strftime("%d/%m/%Y à %H:%M")
            worksheet['B6'] = date_export
            
            worksheet['A3'].font = Font(bold=True, color="1F4E78")
            worksheet['A4'].font = Font(bold=True, color="1F4E78")
            worksheet['A5'].font = Font(bold=True, color="1F4E78")
            worksheet['A6'].font = Font(bold=True, color="1F4E78")
            
            if logo_bytes:
                try:
                    from openpyxl.drawing.image import Image as OpenpyxlImage
                    img = OpenpyxlImage(BytesIO(logo_bytes))
                    # Redimensionner l'image pour qu'elle tienne dans l'en-tête (Hauteur ~80px)
                    ratio = 80 / img.height
                    img.height = 80
                    img.width = int(img.width * ratio)
                    worksheet.add_image(img, 'F2')
                except Exception:
                    pass
            
            col_widths = {'A': 20, 'B': 25, 'C': 35, 'D': 30, 'E': 10, 'F': 15, 'G': 18, 'H': 18}
            for col_letter, width in col_widths.items():
                worksheet.column_dimensions[col_letter].width = width
                
            for col_num in range(len(df.columns)):
                cell = worksheet.cell(row=9, column=col_num+1)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_idx, row in enumerate(df.values, 10):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if col_idx in [7, 8]: cell.number_format = '#,##0.00'
                        
            total_row = len(df) + 10
            worksheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
            worksheet.cell(row=total_row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
            worksheet.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
            
            cell_total = worksheet.cell(row=total_row, column=8, value=total_general)
            cell_total.font = Font(bold=True, color="9C0006")
            cell_total.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell_total.number_format = '#,##0.00'
            cell_total.border = border

        return output.getvalue()

    @staticmethod
    def to_csv(df):
        return df.to_csv(index=False, sep=';').encode('utf-8-sig')


# ==========================================
# INTERFACE LOGICIEL
# ==========================================

# Affichage du Header PRO
st.markdown("""
<div class="header-container">
    <div class="logo-icon">🏗️</div>
    <div>
        <p class="app-title">METRE-TEST SYSTEM</p>
        <p class="app-subtitle">Solution Automatisée d'Extraction de Quantités et Métré BTP</p>
    </div>
</div>
""", unsafe_allow_html=True)

# Initialisation de la mémoire du navigateur (Session State) pour sauvegarder les résultats
if "df" not in st.session_state: st.session_state.df = None
if "total_general" not in st.session_state: st.session_state.total_general = 0.0
if "last_file_name" not in st.session_state: st.session_state.last_file_name = None

if "metadata" not in st.session_state: st.session_state.metadata = {}
if "logo_bytes" not in st.session_state: st.session_state.logo_bytes = None
if "plan_preview" not in st.session_state: st.session_state.plan_preview = None

col1, col2 = st.columns([1, 2])
with col1:
    uploaded_file = st.file_uploader("Étape 1 : Importer le Plan (Format PDF)", type=["pdf"])

if uploaded_file is not None:
    # Si l'utilisateur importe un nouveau fichier, on réinitialise les résultats
    if st.session_state.last_file_name != uploaded_file.name:
        st.session_state.df = None
        st.session_state.total_general = 0.0
        st.session_state.metadata = {}
        st.session_state.logo_bytes = None
        st.session_state.plan_preview = None
        st.session_state.last_file_name = uploaded_file.name
        
    with col1:
        # Ajout d'un bouton pour lancer l'analyse manuellement
        start_btn = st.button("🚀 Étape 2 : Lancer l'Analyse Automatique", type="primary", use_container_width=True)
        
    if start_btn:
        doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
        
        # Génération de l'aperçu du plan (Première page)
        try:
            pix = doc[0].get_pixmap(dpi=150)
            st.session_state.plan_preview = Image.open(io.BytesIO(pix.tobytes("jpeg")))
        except Exception:
            pass
        
        # Extraction du logo (la plus grande image trouvée dans le document)
        logo_bytes = None
        for page in doc:
            images = page.get_images(full=True)
            if images:
                max_size = 0
                for img in images:
                    try:
                        base_image = doc.extract_image(img[0])
                        w, h = base_image["width"], base_image["height"]
                        if w * h > max_size:
                            max_size = w * h
                            best_img = base_image["image"]
                    except: pass
                if max_size > 5000: # Ignorer les petites icônes
                    logo_bytes = best_img
                    break
        st.session_state.logo_bytes = logo_bytes
        
        # Extraction hybride du texte complet et des images pour la Vision IA
        text, images_b64 = ExtractionEngine.extract_all(doc)
        
        if text and len(text.strip()) > 10:
            with st.spinner("⏳ Analyse intelligente avec Vision IA en cours... Veuillez patienter (Cela peut prendre 10 à 30 secondes)."):
                
                # Exécution silencieuse et automatique de l'analyse intelligente
                resultats_dict = ParserMetier.parse_with_ai(text, images_b64)
                metadata = resultats_dict.get("metadata", {})
                resultats = resultats_dict.get("materiaux", [])
                raw_response = resultats_dict.get("raw_response", "")
                
                if len(resultats) > 0:
                    grouped_data = {}
                    tot = 0.0
                    
                    for item in resultats:
                        role = item.get("role", "Divers")
                        ref = item.get("element", "Inconnu")
                        info_db = get_item_info(ref)
                        
                        unite = item.get("unite", "")
                        if unite == "": 
                            unite = info_db["unite"]
                            
                        nbre_pieces = item.get("nbre_pieces", 1)
                        if nbre_pieces is None or str(nbre_pieces).strip() == "": nbre_pieces = 1
                        else:
                            try: nbre_pieces = int(nbre_pieces)
                            except: nbre_pieces = 1
                        
                        longueur = item.get("longueur_unitaire_m", None)
                        poids_ml_or_u = info_db["poids_u"]
                        
                        if unite == "ml" and longueur is not None and str(longueur).replace('.','',1).isdigit():
                            l_m = float(longueur)
                            long_mm = int(l_m * 1000)
                            poids_kg_unt = l_m * poids_ml_or_u
                            poids_kg_m = poids_ml_or_u
                        else:
                            long_mm = "------"
                            poids_kg_unt = poids_ml_or_u
                            poids_kg_m = "------" if unite == "U" else poids_ml_or_u
                            
                        poids_tot_kg = nbre_pieces * poids_kg_unt
                        
                        infos_supp = item.get("infos", "")
                        key = f"{role}____{ref}____{infos_supp}____{long_mm}____{poids_kg_m}____{poids_kg_unt}"
                        
                        if key in grouped_data:
                            grouped_data[key]["Quantité"] += nbre_pieces
                            grouped_data[key]["Poids Tot Kg"] += poids_tot_kg
                        else:
                            grouped_data[key] = {
                                "Nomenclatures": role,
                                "Quantité": nbre_pieces,
                                "Désignation": ref,
                                "Long (mm)": long_mm,
                                "Poids Kg/(m)": poids_kg_m,
                                "Poids Kg/Unt": round(poids_kg_unt, 3),
                                "Poids Tot Kg": poids_tot_kg
                            }
                        
                        tot += poids_tot_kg
                        
                    data = list(grouped_data.values())
                        
                    # Sauvegarde dans la session (Mémoire)
                    st.session_state.df = pd.DataFrame(data).sort_values(by="Poids Tot Kg", ascending=False)
                    st.session_state.total_general = tot
                    st.session_state.metadata = metadata
                else:
                    st.warning("⚠️ Aucun élément trouvé dans ce plan.")
                    with st.expander("🔍 Mode Débogage (Voir pourquoi l'IA n'a rien trouvé)"):
                        st.write("Ceci arrive souvent si le plan est une image (Scanné) ou s'il ne contient pas de vrai texte.")
                        st.text_area("Texte extrait du PDF (Ce que l'IA a vu) :", text[:2000], height=200)
                        st.text_area("Réponse brute de l'IA :", raw_response, height=200)
        else:
            st.error("❌ Ce PDF est une image complète (Scanné) sans texte vectoriel. L'OCR est obligatoire pour l'analyser.")

    # Affichage des résultats s'ils sont dans la mémoire
    if st.session_state.df is not None:
        df = st.session_state.df
        total_general = st.session_state.total_general
        metadata = st.session_state.metadata
        
        with col2:
            st.metric(label="TOTAL GÉNÉRAL", value=f"{total_general:,.2f} KG")
            
            st.write("---")
            st.write("### 📤 Exporter le Métré")
            file_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
            
            st.download_button("📊 Télécharger Fichier EXCEL (TEST)", Exporter.to_excel(df, total_general, metadata, st.session_state.logo_bytes), f"METRE_{file_date}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            st.download_button("📑 Télécharger Données (CSV)", Exporter.to_csv(df), f"METRE_{file_date}.csv", "text/csv", use_container_width=True)
        
        st.write("### 📋 Étape 3 : Résultat du Métré")
        
        # Affichage des métadonnées
        md_col1, md_col2, md_col3 = st.columns(3)
        md_col1.info(f"**🏢 Projet :** {metadata.get('projet', 'Non spécifié')}")
        md_col2.info(f"**💼 Client/Bureau :** {metadata.get('societe', 'Non spécifié')}")
        
        date_plan = metadata.get('date_plan', '')
        date_export = datetime.datetime.now().strftime("%d/%m/%Y à %H:%M")
        md_col3.info(f"**📅 Date / Heure :** {date_export}")
        
        # Création d'une copie du dataframe pour l'affichage avec la ligne TOTAL
        df_display = df.copy()
        total_row_df = pd.DataFrame([{
            "Nomenclatures": "TOTAL GÉNÉRAL", "Quantité": None, "Désignation": "", "Long (mm)": "", 
            "Poids Kg/(m)": "", "Poids Kg/Unt": None, "Poids Tot Kg": total_general
        }])
        df_display = pd.concat([df_display, total_row_df], ignore_index=True)
        
        # Application d'un style spécifique pour la ligne Total
        def highlight_total(s):
            if s.name == len(df_display) - 1: return ['background-color: #f39c12; color: white; font-weight: bold'] * len(s)
            return [''] * len(s)
            
        st.dataframe(df_display.style.apply(highlight_total, axis=1).format({"Poids Kg/Unt": "{:,.3f}", "Poids Tot Kg": "{:,.2f}"}, na_rep=""), use_container_width=True)
        
        st.write("### 📝 Synthèse du Plan")
        st.info(metadata.get('description', "Aucune description trouvée dans ce plan."))
        
        if st.session_state.plan_preview:
            st.write("### 🖼️ Aperçu du Plan")
            st.image(st.session_state.plan_preview, use_container_width=True)
